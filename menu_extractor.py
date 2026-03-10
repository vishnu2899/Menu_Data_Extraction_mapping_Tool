import json
import pytesseract
from PIL import Image
import openpyxl
import pandas as pd
import os

# Set the path to tesseract executable
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def extract_text_from_image(image_path):
    """Extract text from image using OCR."""
    try:
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        return text
    except Exception as e:
        print(f"Error extracting text from {image_path}: {e}")
        return ""

def parse_menu_text(text):
    """Parse the extracted text to identify menu categories, items, prices, etc."""
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    menu_data = {
        'categories': [],
        'items': []
    }
    current_category = None
    accumulated_item = []
    for line in lines:
        if line.isupper() and len(line) > 2:
            # Likely a category
            current_category = line
            menu_data['categories'].append({'name': line})
        else:
            # Accumulate until we find a price
            accumulated_item.append(line)
            # Check if this line or accumulated has a price (digits)
            full_item = ' '.join(accumulated_item)
            # Find the last part that looks like price
            parts = full_item.split()
            price = None
            name_desc = full_item
            for i in range(len(parts) - 1, -1, -1):
                if parts[i].replace('/', '').replace('-', '').replace('.', '').replace(',', '').isdigit():
                    price = ' '.join(parts[i:])
                    name_desc = ' '.join(parts[:i])
                    break
            if price:
                # Parse name and description
                # Assume description in parentheses
                import re
                match = re.match(r'(.+?)\s*\((.+)\)', name_desc)
                if match:
                    name = match.group(1).strip()
                    description = match.group(2).strip()
                else:
                    name = name_desc
                    description = ''
                # Handle variations if / in price
                if '/' in price:
                    variations = []
                    var_parts = price.split('/')
                    for vp in var_parts:
                        vp = vp.strip()
                        if vp:
                            variations.append({'price': vp})
                    price = None  # Set to None if variations
                else:
                    variations = None
                menu_data['items'].append({
                    'name': name,
                    'price': price,
                    'variations': variations,
                    'category': current_category,
                    'description': description
                })
                accumulated_item = []  # Reset
    return menu_data

def load_json_reference(json_path):
    """Load the JSON reference file."""
    with open(json_path, 'r') as f:
        return json.load(f)

def map_data(extracted_data, json_ref):
    """Map extracted data to JSON reference fields, including all extracted data."""
    mapped = {
        'restaurant': json_ref['restaurants'][0]['details'],
        'areas': json_ref['areas'],
        'categories': [],
        'items': []
    }
    # For categories, include all extracted, merge with JSON if match
    for cat in extracted_data['categories']:
        cat_entry = {'categoryname': cat['name'], 'categoryid': '', 'active': '', 'categoryrank': ''}
        for ref_cat in json_ref['categories']:
            if ref_cat['categoryname'].lower() == cat['name'].lower():
                cat_entry.update(ref_cat)
                break
        mapped['categories'].append(cat_entry)
    # For items, include all extracted, merge with JSON if match
    for item in extracted_data['items']:
        item_entry = {
            'itemname': item['name'],
            'itemdescription': item.get('description', ''),
            'price': item.get('price', ''),
            'itemrank': '',
            'instock': '',
            'item_image_url': '',
            'variation': item.get('variations', []),
            'addon': []
        }
        for ref_item in json_ref['items']:
            if ref_item['itemname'].lower() == item['name'].lower():
                item_entry.update(ref_item)
                # Update price if extracted has it
                if item.get('price'):
                    item_entry['price'] = item['price']
                break
        mapped['items'].append(item_entry)
    return mapped

def save_to_excel(mapped_data, template_path, output_path):
    """Format mapped data into the provided Excel template with all columns."""
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading template: {e}. Creating new workbook with template columns.")
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add headers if creating new
        headers = ['restaurant_name', 'area_id', 'area_display_name', 'category_id', 'category_name', 'category_image_url', 'category_timings', 'category_rank', 'item_id', 'item_name', 'item_description', 'price', 'rank', 'category_id.1', 'image_url', 'instock', 'variation_item_id', 'variation_id', 'variation_name', 'variation_price', 'addon_name', 'addon_item_selection', 'addon_item_selection_min', 'addon_item_selection_max', 'addon_price', 'addon_id', 'addon_group_id', 'addon_group_name']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    row = ws.max_row + 1
    # For each item
    for item in mapped_data['items']:
        cat = None
        for c in mapped_data['categories']:
            if c.get('categoryid') == item.get('item_categoryid'):
                cat = c
                break
        if not cat:
            cat = {'categoryid': '', 'categoryname': '', 'category_image_url': '', 'categorytimings': '', 'categoryrank': ''}
        ws.cell(row=row, column=1, value=mapped_data['restaurant']['restaurantname'])  # restaurant_name
        ws.cell(row=row, column=2, value=mapped_data['areas'][0]['areaid'] if mapped_data['areas'] else '')  # area_id
        ws.cell(row=row, column=3, value=mapped_data['areas'][0]['displayname'] if mapped_data['areas'] else '')  # area_display_name
        ws.cell(row=row, column=4, value=cat.get('categoryid', ''))  # category_id
        ws.cell(row=row, column=5, value=cat.get('categoryname', ''))  # category_name
        ws.cell(row=row, column=6, value=cat.get('category_image_url', ''))  # category_image_url
        ws.cell(row=row, column=7, value=cat.get('categorytimings', ''))  # category_timings
        ws.cell(row=row, column=8, value=cat.get('categoryrank', ''))  # category_rank
        ws.cell(row=row, column=9, value=item.get('itemid', ''))  # item_id
        ws.cell(row=row, column=10, value=item.get('itemname', ''))  # item_name
        ws.cell(row=row, column=11, value=item.get('itemdescription', ''))  # item_description
        ws.cell(row=row, column=12, value=item.get('price', ''))  # price
        ws.cell(row=row, column=13, value=item.get('itemrank', ''))  # rank
        ws.cell(row=row, column=14, value=cat.get('categoryid', ''))  # category_id.1
        ws.cell(row=row, column=15, value=item.get('item_image_url', ''))  # image_url
        ws.cell(row=row, column=16, value=item.get('instock', ''))  # instock
        # Variations
        variations = item.get('variation', [])
        if variations:
            var_names = '; '.join([str(v.get('name', '')) for v in variations])
            var_prices = '; '.join([str(v.get('price', '')) for v in variations])
            var_ids = '; '.join([str(v.get('variationid', '')) for v in variations])
            var_item_ids = '; '.join([str(v.get('id', '')) for v in variations])
        else:
            var_names = var_prices = var_ids = var_item_ids = ''
        ws.cell(row=row, column=17, value=var_item_ids)  # variation_item_id
        ws.cell(row=row, column=18, value=var_ids)  # variation_id
        ws.cell(row=row, column=19, value=var_names)  # variation_name
        ws.cell(row=row, column=20, value=var_prices)  # variation_price
        # Add-ons
        addons = item.get('addon', [])
        if addons:
            addon_names = '; '.join([str(a.get('addon_name', '')) for a in addons])
            addon_selections = '; '.join([str(a.get('addon_item_selection', '')) for a in addons])
            addon_mins = '; '.join([str(a.get('addon_item_selection_min', '')) for a in addons])
            addon_maxs = '; '.join([str(a.get('addon_item_selection_max', '')) for a in addons])
            addon_prices = '; '.join([str(a.get('addon_price', '')) for a in addons])
            addon_ids = '; '.join([str(a.get('addon_id', '')) for a in addons])
            addon_group_ids = '; '.join([str(a.get('addon_group_id', '')) for a in addons])
            addon_group_names = '; '.join([str(a.get('addon_group_name', '')) for a in addons])
        else:
            addon_names = addon_selections = addon_mins = addon_maxs = addon_prices = addon_ids = addon_group_ids = addon_group_names = ''
        ws.cell(row=row, column=21, value=addon_names)  # addon_name
        ws.cell(row=row, column=22, value=addon_selections)  # addon_item_selection
        ws.cell(row=row, column=23, value=addon_mins)  # addon_item_selection_min
        ws.cell(row=row, column=24, value=addon_maxs)  # addon_item_selection_max
        ws.cell(row=row, column=25, value=addon_prices)  # addon_price
        ws.cell(row=row, column=26, value=addon_ids)  # addon_id
        ws.cell(row=row, column=27, value=addon_group_ids)  # addon_group_id
        ws.cell(row=row, column=28, value=addon_group_names)  # addon_group_name
        row += 1
    wb.save(output_path)
    print(f"Process completed. Output saved to {output_path}")

def main():
    json_path = 'data_reference.json'
    template_path = 'Python Dev Task Sample.xlsx'
    output_path = 'output_mapped.xlsx'
    image_paths = ['task_menu_1.png', 'task_menu_2.png']

    json_ref = load_json_reference(json_path)
    all_extracted = {'categories': [], 'items': []}

    for img_path in image_paths:
        text = extract_text_from_image(img_path)
        print(f"Extracted text from {img_path}:\n{text}\n")
        data = parse_menu_text(text)
        all_extracted['categories'].extend(data['categories'])
        all_extracted['items'].extend(data['items'])

    mapped_data = map_data(all_extracted, json_ref)
    save_to_excel(mapped_data, template_path, output_path)
    print(f"Process completed. Output saved to {output_path}")

if __name__ == "__main__":
    main()
